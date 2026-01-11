#!/usr/bin/env python3
"""
Excel Spreadsheet Creator and Editor.

Creates and edits .xlsx files using openpyxl with proper formula support,
formatting, and financial model conventions.

Usage:
    uvx --with openpyxl python xlsx_create.py --spec spec.json --output report.xlsx
    uvx --with openpyxl python xlsx_create.py --action add_sheet --input existing.xlsx \
        --spec sheet_spec.json --output updated.xlsx

Actions:
    create      Create a new spreadsheet (default)
    add_sheet   Add a sheet to existing spreadsheet
    update      Update cells in existing spreadsheet
"""

import json
import sys
import argparse
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side,
        NamedStyle,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:
    print("Error: openpyxl is required. Run with: uvx --with openpyxl python xlsx_create.py ...", file=sys.stderr)
    sys.exit(1)


# Financial model color constants (RGB hex without #)
COLORS = {
    "input": "0070C0",       # Blue - input values
    "formula": "000000",     # Black - formulas
    "link": "00B050",        # Green - cross-sheet links
    "external": "FF0000",    # Red - external references
    "assumption": "FFFF00",  # Yellow background - key assumptions
    "header": "4472C4",      # Medium blue - headers
    "header_text": "FFFFFF", # White - header text
}


def hex_to_rgb(hex_color: str) -> str:
    """Convert hex color (with or without #) to ARGB format for openpyxl."""
    color = hex_color.lstrip("#")
    # Ensure 6 characters
    if len(color) == 3:
        color = "".join([c * 2 for c in color])
    return color.upper()


def parse_cell_range(range_str: str) -> tuple[str, str]:
    """
    Parse a cell range like 'A1:B5' into start and end cells.

    Returns:
        Tuple of (start_cell, end_cell)
    """
    if ":" in range_str:
        start, end = range_str.split(":")
        return start.strip(), end.strip()
    return range_str.strip(), range_str.strip()


def apply_formatting(ws, cell_range: str, formatting: dict[str, Any]) -> None:
    """
    Apply formatting to a range of cells.

    Args:
        ws: Worksheet
        cell_range: Cell range like 'A1:B5' or 'A1'
        formatting: Dictionary of formatting options
    """
    start_cell, end_cell = parse_cell_range(cell_range)

    # Parse coordinates
    start_col, start_row = coordinate_from_string(start_cell)
    end_col, end_row = coordinate_from_string(end_cell)
    start_col_idx = column_index_from_string(start_col)
    end_col_idx = column_index_from_string(end_col)

    # Build font
    font_kwargs = {}
    if formatting.get("bold"):
        font_kwargs["bold"] = True
    if formatting.get("italic"):
        font_kwargs["italic"] = True
    if formatting.get("fontColor"):
        font_kwargs["color"] = hex_to_rgb(formatting["fontColor"])
    if formatting.get("fontSize"):
        font_kwargs["size"] = formatting["fontSize"]
    if formatting.get("fontName"):
        font_kwargs["name"] = formatting["fontName"]

    font = Font(**font_kwargs) if font_kwargs else None

    # Build fill
    fill = None
    if formatting.get("fill"):
        fill = PatternFill(
            start_color=hex_to_rgb(formatting["fill"]),
            end_color=hex_to_rgb(formatting["fill"]),
            fill_type="solid",
        )

    # Build alignment
    alignment = None
    if formatting.get("alignment") or formatting.get("wrap"):
        alignment = Alignment(
            horizontal=formatting.get("alignment", "general"),
            wrap_text=formatting.get("wrap", False),
        )

    # Build border
    border = None
    if formatting.get("border"):
        side = Side(style="thin", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)

    # Number format
    number_format = formatting.get("numberFormat")

    # Apply to all cells in range
    for row in range(start_row, end_row + 1):
        for col in range(start_col_idx, end_col_idx + 1):
            cell = ws.cell(row=row, column=col)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if number_format:
                cell.number_format = number_format


def is_formula(value: Any) -> bool:
    """Check if a value is an Excel formula."""
    return isinstance(value, str) and value.startswith("=")


def create_sheet(ws, sheet_spec: dict[str, Any]) -> None:
    """
    Populate a worksheet from a specification.

    Args:
        ws: Worksheet to populate
        sheet_spec: Sheet specification dictionary
    """
    # Set sheet name if specified
    if "name" in sheet_spec:
        ws.title = sheet_spec["name"]

    # Write data
    data = sheet_spec.get("data", [])
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            if cell_value is None:
                continue
            elif is_formula(cell_value):
                # Write formula
                cell.value = cell_value
                # Apply formula color (black)
                cell.font = Font(color=COLORS["formula"])
            elif isinstance(cell_value, (int, float)):
                cell.value = cell_value
            else:
                cell.value = str(cell_value)

    # Set column widths
    column_widths = sheet_spec.get("columnWidths", {})
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter.upper()].width = width

    # Set row heights
    row_heights = sheet_spec.get("rowHeights", {})
    for row_num, height in row_heights.items():
        ws.row_dimensions[int(row_num)].height = height

    # Apply formatting
    formatting = sheet_spec.get("formatting", {})
    for cell_range, format_spec in formatting.items():
        apply_formatting(ws, cell_range, format_spec)

    # Freeze panes
    if "freezePanes" in sheet_spec:
        ws.freeze_panes = sheet_spec["freezePanes"]

    # Auto-filter
    if "autoFilter" in sheet_spec:
        ws.auto_filter.ref = sheet_spec["autoFilter"]


def create_workbook(spec: dict[str, Any], output_path: str) -> None:
    """
    Create a new workbook from a specification.

    Args:
        spec: Workbook specification
        output_path: Output file path
    """
    wb = Workbook()

    # Remove default sheet if we're creating our own
    sheets = spec.get("sheets", [])
    if sheets:
        # Use the first sheet as the default
        default_sheet = wb.active
        create_sheet(default_sheet, sheets[0])

        # Create additional sheets
        for sheet_spec in sheets[1:]:
            ws = wb.create_sheet()
            create_sheet(ws, sheet_spec)
    else:
        # No sheets specified, create an empty workbook
        pass

    # Set workbook properties
    properties = spec.get("properties", {})
    if properties.get("title"):
        wb.properties.title = properties["title"]
    if properties.get("creator"):
        wb.properties.creator = properties["creator"]

    # Save
    wb.save(output_path)
    print(f"Created: {output_path}")


def add_sheet_to_workbook(
    input_path: str,
    sheet_spec: dict[str, Any],
    output_path: str,
) -> None:
    """
    Add a sheet to an existing workbook.

    Args:
        input_path: Path to existing workbook
        sheet_spec: Sheet specification
        output_path: Output file path
    """
    wb = load_workbook(input_path)

    # Create new sheet
    sheet_name = sheet_spec.get("name", "Sheet")

    # Handle duplicate names
    if sheet_name in wb.sheetnames:
        base_name = sheet_name
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{base_name}_{counter}"
            counter += 1
        sheet_spec["name"] = sheet_name

    ws = wb.create_sheet(title=sheet_name)
    create_sheet(ws, sheet_spec)

    wb.save(output_path)
    print(f"Added sheet '{sheet_name}' to: {output_path}")


def update_cells(
    input_path: str,
    updates: dict[str, Any],
    output_path: str,
) -> None:
    """
    Update cells in an existing workbook.

    Args:
        input_path: Path to existing workbook
        updates: Dictionary of {sheet_name: {cell_ref: value, ...}, ...}
        output_path: Output file path
    """
    wb = load_workbook(input_path)

    for sheet_name, cells in updates.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found", file=sys.stderr)
            continue

        ws = wb[sheet_name]
        for cell_ref, value in cells.items():
            ws[cell_ref] = value

    wb.save(output_path)
    print(f"Updated: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Create and edit Excel spreadsheets with formula support",
        epilog="Example: python xlsx_create.py --spec spec.json --output report.xlsx",
    )
    parser.add_argument(
        "--action",
        "-a",
        choices=["create", "add_sheet", "update"],
        default="create",
        help="Action to perform (default: create)",
    )
    parser.add_argument(
        "--spec",
        "-s",
        help="Path to JSON specification file",
    )
    parser.add_argument(
        "--input",
        "-i",
        help="Input .xlsx file (for add_sheet/update actions)",
    )
    parser.add_argument(
        "--output",
        "-o",
        required=True,
        help="Output .xlsx file path",
    )
    parser.add_argument(
        "--stdin",
        action="store_true",
        help="Read specification from stdin",
    )

    args = parser.parse_args()

    # Read specification
    if args.stdin:
        spec_json = sys.stdin.read()
    elif args.spec:
        if not Path(args.spec).exists():
            print(f"Error: Spec file not found: {args.spec}", file=sys.stderr)
            sys.exit(1)
        with open(args.spec) as f:
            spec_json = f.read()
    else:
        print("Error: Either --spec or --stdin is required", file=sys.stderr)
        sys.exit(1)

    try:
        spec = json.loads(spec_json)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON: {e}", file=sys.stderr)
        sys.exit(1)

    # Perform action
    try:
        if args.action == "create":
            create_workbook(spec, args.output)

        elif args.action == "add_sheet":
            if not args.input:
                print("Error: --input is required for add_sheet action", file=sys.stderr)
                sys.exit(1)
            if not Path(args.input).exists():
                print(f"Error: Input file not found: {args.input}", file=sys.stderr)
                sys.exit(1)
            add_sheet_to_workbook(args.input, spec, args.output)

        elif args.action == "update":
            if not args.input:
                print("Error: --input is required for update action", file=sys.stderr)
                sys.exit(1)
            if not Path(args.input).exists():
                print(f"Error: Input file not found: {args.input}", file=sys.stderr)
                sys.exit(1)
            update_cells(args.input, spec, args.output)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
