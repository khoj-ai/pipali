#!/usr/bin/env python3
"""
Excel Formula Recalculation and Error Detection.

Recalculates all formulas in an Excel file using LibreOffice (if available)
and scans for Excel errors (#REF!, #DIV/0!, etc.).

Usage:
    uvx --with openpyxl python xlsx_recalc.py report.xlsx
    uvx --with openpyxl python xlsx_recalc.py report.xlsx --timeout 60

Requirements:
    - openpyxl (for reading results)
    - LibreOffice (optional, for recalculation)

If LibreOffice is not installed, the script will only scan for existing errors
without recalculating formulas.
"""

import json
import os
import platform
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl is required. Run with: uvx --with openpyxl python xlsx_recalc.py ...", file=sys.stderr)
    sys.exit(1)


# Excel error values to detect
EXCEL_ERRORS = [
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
]


def find_libreoffice() -> str | None:
    """
    Find the LibreOffice executable.

    Returns:
        Path to soffice executable, or None if not found
    """
    system = platform.system()

    if system == "Darwin":  # macOS
        paths = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            shutil.which("soffice"),
        ]
    elif system == "Windows":
        paths = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
            shutil.which("soffice"),
        ]
    else:  # Linux
        paths = [
            "/usr/bin/soffice",
            "/usr/local/bin/soffice",
            shutil.which("soffice"),
            shutil.which("libreoffice"),
        ]

    for path in paths:
        if path and Path(path).exists():
            return path

    return None


def setup_libreoffice_macro(soffice: str) -> bool:
    """
    Setup LibreOffice macro for recalculation.

    Args:
        soffice: Path to soffice executable

    Returns:
        True if setup successful
    """
    system = platform.system()

    if system == "Darwin":
        macro_dir = Path.home() / "Library/Application Support/LibreOffice/4/user/basic/Standard"
    elif system == "Windows":
        macro_dir = Path(os.environ.get("APPDATA", "")) / "LibreOffice/4/user/basic/Standard"
    else:
        macro_dir = Path.home() / ".config/libreoffice/4/user/basic/Standard"

    macro_file = macro_dir / "Module1.xba"

    # Check if macro already exists
    if macro_file.exists():
        content = macro_file.read_text()
        if "RecalculateAndSave" in content:
            return True

    # Initialize LibreOffice to create directories
    if not macro_dir.exists():
        try:
            subprocess.run(
                [soffice, "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=30,
            )
        except subprocess.TimeoutExpired:
            pass

        macro_dir.mkdir(parents=True, exist_ok=True)

    # Write macro
    macro_content = '''<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>'''

    try:
        macro_file.write_text(macro_content)
        return True
    except Exception as e:
        print(f"Warning: Could not setup macro: {e}", file=sys.stderr)
        return False


def recalculate_with_libreoffice(xlsx_path: str, soffice: str, timeout: int) -> bool:
    """
    Recalculate formulas using LibreOffice.

    Args:
        xlsx_path: Path to Excel file
        soffice: Path to soffice executable
        timeout: Timeout in seconds

    Returns:
        True if recalculation successful
    """
    if not setup_libreoffice_macro(soffice):
        return False

    abs_path = str(Path(xlsx_path).absolute())

    # Build command
    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "macro:///Standard.Module1.RecalculateAndSave",
        abs_path,
    ]

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        print(f"Warning: LibreOffice timed out after {timeout}s", file=sys.stderr)
        return False
    except Exception as e:
        print(f"Warning: LibreOffice error: {e}", file=sys.stderr)
        return False


def scan_for_errors(xlsx_path: str) -> dict:
    """
    Scan Excel file for formula errors.

    Args:
        xlsx_path: Path to Excel file

    Returns:
        Dictionary with error details
    """
    try:
        # Load with data_only=True to see calculated values
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return {"error": f"Could not load file: {e}"}

    error_details = {err: [] for err in EXCEL_ERRORS}
    total_errors = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    for err in EXCEL_ERRORS:
                        if err in cell.value:
                            location = f"{sheet_name}!{cell.coordinate}"
                            error_details[err].append(location)
                            total_errors += 1
                            break

    wb.close()

    # Count formulas
    try:
        wb_formulas = load_workbook(xlsx_path, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        wb_formulas.close()
    except Exception:
        formula_count = -1

    # Build result
    result = {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "total_formulas": formula_count,
        "error_summary": {},
    }

    for err_type, locations in error_details.items():
        if locations:
            result["error_summary"][err_type] = {
                "count": len(locations),
                "locations": locations[:20],  # Limit to first 20
            }

    return result


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="Recalculate Excel formulas and detect errors",
        epilog="Example: python xlsx_recalc.py report.xlsx",
    )
    parser.add_argument(
        "xlsx_file",
        help="Path to Excel file",
    )
    parser.add_argument(
        "--timeout",
        "-t",
        type=int,
        default=30,
        help="Timeout for LibreOffice recalculation (default: 30s)",
    )
    parser.add_argument(
        "--no-recalc",
        action="store_true",
        help="Skip recalculation, only scan for errors",
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output as JSON",
    )

    args = parser.parse_args()

    # Validate file exists
    if not Path(args.xlsx_file).exists():
        print(f"Error: File not found: {args.xlsx_file}", file=sys.stderr)
        sys.exit(1)

    # Find LibreOffice
    soffice = find_libreoffice()
    recalc_success = False

    if not args.no_recalc:
        if soffice:
            print(f"Found LibreOffice: {soffice}", file=sys.stderr)
            print("Recalculating formulas...", file=sys.stderr)
            recalc_success = recalculate_with_libreoffice(
                args.xlsx_file, soffice, args.timeout
            )
            if recalc_success:
                print("Recalculation complete.", file=sys.stderr)
            else:
                print("Recalculation failed, scanning existing values...", file=sys.stderr)
        else:
            print(
                "LibreOffice not found. Scanning for existing errors only.",
                file=sys.stderr,
            )
            print(
                "Install LibreOffice for formula recalculation:",
                file=sys.stderr,
            )
            print("  macOS: brew install --cask libreoffice", file=sys.stderr)
            print("  Ubuntu: sudo apt install libreoffice", file=sys.stderr)
            print("  Windows: Download from libreoffice.org", file=sys.stderr)

    # Scan for errors
    result = scan_for_errors(args.xlsx_file)
    result["recalculated"] = recalc_success

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        # Human-readable output
        print(f"\nFile: {args.xlsx_file}")
        print(f"Recalculated: {'Yes' if recalc_success else 'No'}")
        print(f"Total formulas: {result.get('total_formulas', 'Unknown')}")
        print(f"Total errors: {result.get('total_errors', 0)}")

        if result.get("error_summary"):
            print("\nErrors found:")
            for err_type, details in result["error_summary"].items():
                print(f"  {err_type}: {details['count']} occurrence(s)")
                for loc in details["locations"][:5]:
                    print(f"    - {loc}")
                if len(details["locations"]) > 5:
                    print(f"    ... and {len(details['locations']) - 5} more")
        else:
            print("\nNo formula errors detected!")


if __name__ == "__main__":
    main()
